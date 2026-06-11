"""Fetch the authoritative SGB-II Jobcenter register from BA statistics.

Source: Gebietsstruktur der Grundsicherungsträger (statistik.arbeitsagentur.de),
updated with every Gebietsstand. Two files:

- Gebietsstruktur-...XLS.xlsx  -> Träger-Nr., name, legal form (gE/zkT), Land
- Zuordnung-Jobcenter.xlsx     -> Träger-Nr. -> Kreisschlüssel (5-digit AGS),
                                  incl. flag for Kreise split across Jobcenter

This is the *authoritative* source for Jobcenter existence and legal form;
PVOG only enriches contacts where it knows the office (gE are missing there).

Output (data/snapshots/ba/<date>/): traeger.jsonl, zuordnung.jsonl
"""
from __future__ import annotations

import sys

import openpyxl

from common import Http, snapshot_dir, write_jsonl

BASE = ("https://statistik.arbeitsagentur.de/DE/Statischer-Content/"
        "Grundlagen/Klassifikationen/Regionale-Gliederungen/")
URL_STRUKTUR = (BASE + "Gebietsstruktur-Traeger-Grundsicherung/"
                "Generische-Publikationen/"
                "Gebietsstruktur-Traeger-Grundsicherung-XLS.xlsx")
URL_ZUORDNUNG = BASE + "Generische-Publikationen/Zuordnung-Jobcenter.xlsx"


def header_index(rows: list[tuple], *needles: str) -> tuple[int, dict[str, int]]:
    """Find the header row containing all needles; return (row_idx, col_map)."""
    for i, row in enumerate(rows):
        cells = [str(c or "").lower() for c in row]
        joined = " ".join(cells)
        if all(n.lower() in joined for n in needles):
            return i, {str(c or "").strip(): j for j, c in enumerate(row)}
    raise ValueError(f"header with {needles} not found")


def col(cols: dict[str, int], needle: str) -> int:
    for name, j in cols.items():
        if needle.lower() in name.lower():
            return j
    raise ValueError(f"column ~'{needle}' not found in {list(cols)}")


def load_sheet(path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [tuple(r) for r in ws.iter_rows(values_only=True)]


def main() -> int:
    http = Http(delay=0.2)
    out = snapshot_dir("ba")

    for url, fname in ((URL_STRUKTUR, "struktur.xlsx"),
                       (URL_ZUORDNUNG, "zuordnung.xlsx")):
        r = http.get(url, timeout=60)
        r.raise_for_status()
        (out / fname).write_bytes(r.content)

    # --- Träger (legal form) -------------------------------------------
    rows = load_sheet(str(out / "struktur.xlsx"))
    hi, cols = header_index(rows, "Typ", "Träger-Nr")
    c_typ, c_nr = col(cols, "Typ"), col(cols, "Träger-Nr")
    c_name, c_land = col(cols, "Jobcenterbezirk"), col(cols, "Land")
    traeger = []
    for row in rows[hi + 1:]:
        typ = str(row[c_typ] or "").strip()
        if typ not in ("gE", "zkT"):
            continue
        traeger.append({
            "traeger_nr": str(row[c_nr] or "").strip(),
            "name": str(row[c_name] or "").strip(),
            "legal_form": typ,
            "land": str(row[c_land] or "").strip(),
        })
    write_jsonl(out / "traeger.jsonl", traeger)
    ge = sum(1 for t in traeger if t["legal_form"] == "gE")
    print(f"traeger: {len(traeger)} ({ge} gE, {len(traeger) - ge} zkT)")

    # --- Zuordnung Träger -> Kreis --------------------------------------
    rows = load_sheet(str(out / "zuordnung.xlsx"))
    hi, cols = header_index(rows, "Träger-Nr", "Kreisschlüssel")
    c_nr = col(cols, "Träger-Nr")
    c_kreis = col(cols, "Kreisschlüssel")
    c_kname = col(cols, "Kreis")
    try:
        c_split = col(cols, "mehrere")
    except ValueError:
        c_split = None
    zuordnung = []
    for row in rows[hi + 1:]:
        nr = str(row[c_nr] or "").strip()
        kreis = str(row[c_kreis] or "").strip()
        if not nr or not kreis or not kreis[:1].isdigit():
            continue
        zuordnung.append({
            "traeger_nr": nr,
            "kreis_ags": kreis.zfill(5),
            "kreis_name": str(row[c_kname] or "").strip(),
            "split": bool(str(row[c_split] or "").strip()) if c_split is not None
                     else False,
        })
    write_jsonl(out / "zuordnung.jsonl", zuordnung)
    print(f"zuordnung: {len(zuordnung)} Träger-Kreis rows, "
          f"{sum(1 for z in zuordnung if z['split'])} split-Kreis rows")
    print(f"snapshot -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
