"""Fetch the Destatis Gemeindeverzeichnis (AuszugGV, quarterly XLSX) and
extract the real 12-digit ARS per Gemeinde.

The ARS (Amtlicher Regionalschlüssel) is Land(2)+RB(1)+Kreis(2)+Verband(4)+
Gemeinde(3); the AGS is the same without the Verband part. The Verband
digits cannot be derived from the AGS, hence this authoritative source.

Falls back to the v1 copy in bac/MIN.REG.AMT/ when the download fails
(Destatis occasionally moves the file).

Output (data/snapshots/destatis/<date>/): gemeinden.jsonl
    {ags, ars, name, plz, satzart}
"""
from __future__ import annotations

import sys

import openpyxl

from common import Http, ROOT, snapshot_dir, write_jsonl

URL = ("https://www.destatis.de/DE/Themen/Laender-Regionen/Regionales/"
       "Gemeindeverzeichnis/Administrativ/Archiv/GVAuszugQ/"
       "AuszugGV2QAktuell.xlsx?__blob=publicationFile")
FALLBACK = ROOT / "bac" / "MIN.REG.AMT" / "AuszugGV2QAktuell.xlsx"


def norm(v) -> str:
    return str(v).strip() if v is not None else ""


def main() -> int:
    out = snapshot_dir("destatis")
    xlsx = out / "AuszugGV.xlsx"
    try:
        r = Http(delay=0.2).get(URL, timeout=90)
        r.raise_for_status()
        if len(r.content) < 100_000:
            raise ValueError(f"suspiciously small file ({len(r.content)}B)")
        xlsx.write_bytes(r.content)
        print(f"downloaded fresh AuszugGV ({len(r.content)} bytes)")
    except Exception as exc:                    # noqa: BLE001
        if not FALLBACK.exists():
            print(f"download failed ({exc}) and no fallback", file=sys.stderr)
            return 1
        xlsx.write_bytes(FALLBACK.read_bytes())
        print(f"download failed ({exc}); using v1 fallback copy "
              f"(older Gebietsstand!)")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "Online" in s), wb.sheetnames[0])
    ws = wb[sheet]

    # The ARS columns sit under one merged "Amtlicher Regionalschlüssel"
    # header without sub-labels, so name-based detection is impossible.
    # The AuszugGV layout has been stable for years:
    #   0=Satzart 1=Textkennzeichen 2=Land 3=RB 4=Kreis 5=VB 6=Gem
    #   7=Gemeindename 13=PLZ
    C_SATZ, C_LAND, C_RB, C_KREIS, C_VB, C_GEM, C_NAME, C_PLZ = \
        0, 2, 3, 4, 5, 6, 7, 13

    gemeinden = []
    for row in ws.iter_rows(values_only=True):
        if norm(row[C_SATZ]) != "60":           # 60 = Gemeinde record
            continue
        land = norm(row[C_LAND]).zfill(2)
        rb = norm(row[C_RB]) or "0"
        kreis = norm(row[C_KREIS]).zfill(2)
        vb = norm(row[C_VB]).zfill(4)
        gem = norm(row[C_GEM]).zfill(3)
        if not (land.isdigit() and kreis.isdigit() and gem.isdigit()):
            continue
        plz = norm(row[C_PLZ]) if len(row) > C_PLZ else ""
        if plz and len(plz) == 4:
            plz = "0" + plz
        gemeinden.append({
            "ags": f"{land}{rb}{kreis}{gem}",
            "ars": f"{land}{rb}{kreis}{vb}{gem}",
            "name": norm(row[C_NAME]),
            "plz": plz if plz.isdigit() and len(plz) == 5 else None,
        })

    # self-check against a known fixture before writing anything: the
    # layout assumption above must hold or the run fails loudly
    fixture = {g["ags"]: g["ars"] for g in gemeinden}
    if len(gemeinden) < 10000 or fixture.get("09779194") != "097790194194":
        print(f"LAYOUT CHECK FAILED: rows={len(gemeinden)}, "
              f"Nördlingen ARS={fixture.get('09779194')} "
              f"(expected 097790194194)", file=sys.stderr)
        return 1

    write_jsonl(out / "gemeinden.jsonl", gemeinden)
    print(f"gemeinden: {len(gemeinden)} -> {out} (layout check passed)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
